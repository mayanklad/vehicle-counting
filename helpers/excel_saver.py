import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

current_path = os.getcwd()

def save_excel(detections, file_name=current_path+'/detection-report.xlsx'):

    try:
        print('Appending...')
        wb = load_workbook(file_name)
        page = wb.active

    except:
        print("File doesnt exist...")
        headers = ['Vehicle (Image name)', 'Type', 'Color', 'Year', 'Month', 'Day', 'Hour (24)', 'Minute', 'Second']
        wb = Workbook()
        page = wb.active
        page.title = 'Report'
        page.append(headers) # write the headers to the first line

    finally:
        for info in detections:
            page.append(info)

        wb.save(filename=file_name)
